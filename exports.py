"""
Auto-export to CSV and XLSX after each processing cycle.
"""

import csv
import json
import logging
import os
from pathlib import Path

from config.loader import get_config
import db

log = logging.getLogger(__name__)

COLUMNS = [
    "id", "date", "supplier", "invoice_number",
    "net_amount", "vat_amount", "vat_rate", "vat_breakdown", "total_amount", "currency",
    "file_path", "source", "confidence", "needs_review",
    "vat_flag", "spike_flag", "hash", "created_at",
]


def _format_vat_breakdown(value) -> str:
    """Convert vat_breakdown list to '10%: 90.01 base, 9.00 VAT | 21%: 82.01 base, 17.22 VAT'."""
    if not value:
        return ""
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (json.JSONDecodeError, ValueError):
            return value
    if not isinstance(value, list):
        return str(value)
    parts = []
    for entry in value:
        rate = entry.get("rate", 0)
        base = entry.get("base", 0)
        amount = entry.get("amount", 0)
        parts.append(f"{rate*100:.0f}%: {base:.2f} base, {amount:.2f} IVA")
    return " | ".join(parts)


def _prepare_row(inv: dict) -> dict:
    row = dict(inv)
    row["vat_breakdown"] = _format_vat_breakdown(inv.get("vat_breakdown"))
    return row


def run_exports():
    cfg = get_config()
    export_cfg = cfg.get("exports", {})
    if not export_cfg.get("auto_export", True):
        return

    formats = export_cfg.get("formats", ["csv", "xlsx"])
    export_dir = cfg["paths"].get("exports", str(Path(cfg["paths"]["db"]).parent))
    os.makedirs(export_dir, exist_ok=True)

    invoices, _ = db.get_all_invoices(page=1, page_size=999999)

    if "csv" in formats:
        _export_csv(invoices, os.path.join(export_dir, "facturas.csv"))
    if "xlsx" in formats:
        _export_xlsx(invoices, os.path.join(export_dir, "facturas.xlsx"))


def _export_csv(invoices: list[dict], path: str):
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows([_prepare_row(inv) for inv in invoices])
        log.info("CSV exported: %s", path)
    except Exception as e:
        log.error("CSV export failed: %s", e)


def _export_xlsx(invoices: list[dict], path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A1A2E")

        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, inv in enumerate(invoices, 2):
            row = _prepare_row(inv)
            for col_idx, col_name in enumerate(COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        # Totals row
        if invoices:
            totals_row = len(invoices) + 2
            ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
            amount_cols = {"net_amount": 5, "vat_amount": 6, "total_amount": 8}
            for col_name, col_idx in amount_cols.items():
                col_letter = get_column_letter(col_idx)
                ws.cell(
                    row=totals_row, column=col_idx,
                    value=f"=SUM({col_letter}2:{col_letter}{totals_row - 1})"
                ).font = Font(bold=True)

        # Auto-width
        for col_idx, col_name in enumerate(COLUMNS, 1):
            max_len = max(len(col_name), 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        wb.save(path)
        log.info("XLSX exported: %s", path)
    except ImportError:
        log.warning("openpyxl not installed — skipping XLSX export")
    except Exception as e:
        log.error("XLSX export failed: %s", e)
