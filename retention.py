"""
Legal retention for Spanish SL invoices.

  Tax (AEAT / Ley General Tributaria art. 66):  4 years
  Commercial (Código de Comercio art. 30):       6 years

We keep 6 years as the safe default. At year-end we generate a per-year
archive package (CSV + XLSX) so each fiscal year stands alone for auditors.
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path

from config.loader import get_config
import db

log = logging.getLogger(__name__)

RETENTION_YEARS = 6


def current_year() -> int:
    return datetime.now(timezone.utc).year


def run_yearly_archives():
    """
    Called at the end of each processing cycle.
    Writes output/YYYY/YYYY_archive.csv and YYYY_archive.xlsx for every year
    that has at least one invoice and whose archive file is missing or stale.
    """
    cfg = get_config()
    output_base = cfg["paths"]["output"]

    with db.transaction() as conn:
        years = [
            row[0]
            for row in conn.execute(
                "SELECT DISTINCT substr(date,1,4) FROM invoices "
                "WHERE date IS NOT NULL ORDER BY 1"
            ).fetchall()
            if row[0]
        ]

    for year in years:
        year_dir = os.path.join(output_base, str(year))
        os.makedirs(year_dir, exist_ok=True)
        csv_path  = os.path.join(year_dir, f"{year}_archive.csv")
        xlsx_path = os.path.join(year_dir, f"{year}_archive.xlsx")

        with db.transaction() as conn:
            rows = conn.execute(
                "SELECT * FROM invoices WHERE substr(date,1,4) = ? ORDER BY date",
                (str(year),),
            ).fetchall()

        invoices = [dict(r) for r in rows]
        if not invoices:
            continue

        _write_csv(invoices, csv_path)
        _write_xlsx(invoices, xlsx_path, year)


def get_retention_warnings() -> list[dict]:
    """
    Return invoices whose fiscal year is within 12 months of the retention
    deadline (6 years after the invoice date).
    """
    cutoff_year = current_year() - RETENTION_YEARS + 1
    with db.transaction() as conn:
        rows = conn.execute(
            """SELECT id, date, supplier, invoice_number, file_path
               FROM invoices
               WHERE substr(date,1,4) <= ?
               ORDER BY date""",
            (str(cutoff_year),),
        ).fetchall()
    return [dict(r) for r in rows]


def years_summary() -> list[dict]:
    """Count invoices and totals per fiscal year."""
    with db.transaction() as conn:
        rows = conn.execute(
            """SELECT substr(date,1,4) as year,
                      COUNT(*)          as count,
                      SUM(total_amount) as total
               FROM invoices
               WHERE date IS NOT NULL
               GROUP BY year
               ORDER BY year DESC"""
        ).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _write_csv(invoices: list[dict], path: str):
    import csv
    from exports import COLUMNS
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            w.writeheader()
            w.writerows(invoices)
        log.debug("Yearly CSV: %s", path)
    except Exception as e:
        log.error("Yearly CSV failed (%s): %s", path, e)


def _write_xlsx(invoices: list[dict], path: str, year: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from exports import COLUMNS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(year)

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1A1A2E")
        for ci, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, inv in enumerate(invoices, 2):
            for ci, col in enumerate(COLUMNS, 1):
                ws.cell(row=ri, column=ci, value=inv.get(col))

        # Totals row
        tr = len(invoices) + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        for col_name, ci in (("net_amount", 5), ("vat_amount", 6), ("total_amount", 8)):
            cl = get_column_letter(ci)
            ws.cell(row=tr, column=ci,
                    value=f"=SUM({cl}2:{cl}{tr-1})").font = Font(bold=True)

        # Retention notice in a separate sheet
        note = wb.create_sheet("Retención Legal")
        note["A1"] = f"Facturas del ejercicio {year}"
        note["A2"] = f"Plazo mínimo de conservación: {RETENTION_YEARS} años"
        note["A3"] = f"Conservar hasta: {int(year) + RETENTION_YEARS}"
        note["A4"] = "Base legal: LGT art. 66 (4 años fiscal) | C.Com art. 30 (6 años mercantil)"

        wb.save(path)
        log.debug("Yearly XLSX: %s", path)
    except ImportError:
        log.warning("openpyxl not installed — skipping yearly XLSX")
    except Exception as e:
        log.error("Yearly XLSX failed (%s): %s", path, e)
